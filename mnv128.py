
import cv2
import numpy as np
import openpyxl
from utils import *

'''
    量化标准：图片对应的灰度图的像素值偏离128的均值绝对值与平均偏差的比值，成为亮度系数
    输入：cv2图片对象；判定阈值：亮度系数超过该阈值则判断为亮度异常
    输出：亮度系数，对应的灰度图的像素值偏离128的均值（用于判断亮度异常图片属于过亮或过暗）
'''
def brightnessCoefficient(img):
    # 把图片转换为单通道的灰度图
    gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # 获取形状以及长宽
    img_shape = gray_img.shape
    height, width = img_shape[0], img_shape[1]
    size = gray_img.size
    # 灰度图的直方图
    hist = cv2.calcHist([gray_img], [0], None, [256], [0, 256])
    ma = 0
    # 计算灰度图像素点偏离128的均值
    reduce_matrix = np.full((height, width), 128)   # np.full 构造一个数组，用指定值填充其元素
    shift_value = gray_img - reduce_matrix
    shift_sum = np.sum(shift_value)
    da = shift_sum / size
    # 计算偏离128的平均偏差
    for i in range(256):
        ma += (abs(i-128-da) * hist[i])
    m = abs(ma / size)
    # 亮度系数
    k = abs(da) / m
    return float(k), da

def test():
	# 读取图片
	data_type = 'data1'
	source_path = 'brightness/pic/' + data_type 
	save_path = 'brightness/result/' + data_type
	pic_paths = readAllPictures(source_path)

	result_list = []
	# 评估图片
	for pic_path in pic_paths:
		pic = cv2.imread(pic_path)
		coef, mean = brightnessCoefficient(pic)
		result_list.append((os.path.split(pic_path)[1], coef, mean))

	# 将图片名和对应的结果存放到xls文件中
	workbook = openpyxl.Workbook()
	sheet = workbook.active
	i = 2
	sheet.cell(1, 1, 'pic')
	sheet.cell(1, 2, 'coef')
	sheet.cell(1, 3, 'mean')
	for item in result_list:
		sheet.cell(i, 1, item[0])
		sheet.cell(i, 2, item[1])
		sheet.cell(i, 3, item[2])
		i += 1
	workbook.save(save_path + 'BCGrey.xls')

	# # 将图片名和对应的结果存放到txt文件中
	# f = open(os.path.join(save_path, 'BCGrey.txt'), 'w')
	# f.write('{:6}, {:10}, {:10}\n'.format('pic', 'coef', 'mean'))
	# for item in result_list:
	# 	f.write('{:6}, {:10.3f}, {:10.3f}\n'.format(item[0], item[1], item[2]))
	# f.close()

test()
