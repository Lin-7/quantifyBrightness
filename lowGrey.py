
import cv2
from utils import *
import openpyxl


'''
	量化标准：较暗的像素占整一张图片的比例
	参数：img：cv2的图片对象；lThreshold：超参数--亮度阈值，低于此值判断为低亮度像素
	返回值：低亮度像素所占比例，像素总数
'''
def lowGrey(img,lThreshold=40):
	# 把图片转换为灰度图
	gray_img = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
	# 获取灰度图矩阵的行数和列数
	r,c = gray_img.shape[:2]
	# 获取灰度图像素总数
	piexs_sum = gray_img.size   # 等于r*c

	dark_sum=0    	# 偏暗的像素 初始化为0个
	dark_prop=0	    # 偏暗像素所占比例初始化为0
	
	# 遍历灰度图的所有像素
	for row in gray_img:
		for value in row:
			if value < lThreshold:
				dark_sum+=1
	dark_prop=dark_sum/(piexs_sum)
	return dark_prop, piexs_sum

def test():
	# 读取图片
	data_type = 'data1'
	source_path = 'brightness/pic/' + data_type 
	save_path = 'brightness/result/' + data_type
	pic_paths = readAllPictures(source_path)

	result_list = []
	# 评估图片
	for pic_path in pic_paths:
		pic = cv2.imread(pic_path)
		dark_prop, piexs_sum = lowGrey(pic)
		result_list.append((os.path.split(pic_path)[1], dark_prop, piexs_sum))

	# 将图片名和对应的结果存放到xls文件中
	workbook = openpyxl.Workbook()
	sheet = workbook.active
	i = 2
	sheet.cell(1, 1, 'pic')
	sheet.cell(1, 2, 'dark_prop')
	sheet.cell(1, 3, 'piexs_sum')
	for item in result_list:
		sheet.cell(i, 1, item[0])
		sheet.cell(i, 2, item[1])
		sheet.cell(i, 3, item[2])
		i += 1
	workbook.save(save_path + 'lowGrey.xls')

	# # 将图片名和对应的结果存放到txt文件中
	# f = open(os.path.join(save_path, 'lowGrey.txt'), 'w')
	# f.write('{:6}, {:8}, {:5}\n'.format('pic', 'dark_prop', 'piexs_sum'))
	# for item in result_list:
	# 	f.write('{:6}, {:8.3f}, {:5}\n'.format(item[0], item[1], item[2]))
	# f.close()

test()

