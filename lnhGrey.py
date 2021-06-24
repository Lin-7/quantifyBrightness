
import cv2
import openpyxl
from utils import *

eps = 1

'''
	量化标准：较亮像素数量和较暗的像素数量的比值
	参数：img：cv2的图片对象；lThreshold：超参数--亮度阈值，低于此值判断为低亮度像素；hThreshold：超参数--亮度阈值，高于此值判断为高亮度像素；
	返回值：较亮像素数量和较暗的像素数量的比值，低亮度像素个数，高亮度像素个数，像素总数
'''
def lnhGrey(img,lThreshold=40,hThreshold=216):
	# 把图片转换为灰度图
	gray_img = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
	# 获取灰度图矩阵的行数和列数
	r,c = gray_img.shape[:2]
	# 获取灰度图像素总数
	piexs_sum = gray_img.size   # 等于r*c

	dark_sum=0    	# 偏暗的像素 初始化为0个
	bright_sum=0    # 偏亮的像素 初始化为0个
	
    # 遍历灰度图的所有像素
	for row in gray_img:
		for value in row:
			if value < lThreshold:
				dark_sum+=1
			if value > hThreshold:
				bright_sum += 1
	brightDarkProp = bright_sum/(dark_sum+eps)
	return brightDarkProp, dark_sum, bright_sum, piexs_sum

def test():
	# 读取图片
	data_type = 'data1'
	source_path = 'brightness/pic/' + data_type 
	save_path = 'brightness/result/' + data_type
	pic_paths = readAllPictures(source_path)

	result_list = []
	# 评估图片
	for pic_path in pic_paths:
		pic = cv2.imread(pic_path)
		brightDarkProp, dark_sum, bright_sum, piexs_sum = lnhGrey(pic)
		result_list.append((os.path.split(pic_path)[1], brightDarkProp, dark_sum, bright_sum, piexs_sum))

	# 将图片名和对应的结果存放到xls文件中
	workbook = openpyxl.Workbook()
	sheet = workbook.active
	i = 2
	sheet.cell(1, 1, 'pic')
	sheet.cell(1, 2, 'brightDarkProp')
	sheet.cell(1, 3, 'dark_sum')
	sheet.cell(1, 4, 'bright_sum')
	sheet.cell(1, 5, 'piexs_sum')
	for item in result_list:
		sheet.cell(i, 1, item[0])
		sheet.cell(i, 2, item[1])
		sheet.cell(i, 3, item[2])
		sheet.cell(i, 4, item[3])
		sheet.cell(i, 5, item[4])
		i += 1
	workbook.save(save_path + 'lnhGrey.xls')

	# # 将图片名和对应的结果存放到txt文件中
	# f = open(os.path.join(save_path, 'lnhGrey.txt'), 'w')
	# f.write('{:6}, {:10}, {:8}, {:8}, {:8}\n'.format('pic', 'brightDarkProp', 'dark_sum', 'bright_sum', 'piexs_sum'))
	# for item in result_list:
	# 	f.write('{:6}, {:10.3f}, {:8}, {:8}, {:8}\n'.format(item[0], item[1], item[2], item[3], item[4]))
	# f.close()

test()
